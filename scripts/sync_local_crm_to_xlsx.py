import json
import os
import sys
from collections import defaultdict
from pathlib import Path

import openpyxl


JSON_SHEETS = [
    "Locations", "Amenities", "Properties", "Leads", "FollowUps",
    "Appointments", "Owners", "Deals", "Tenancies", "Logs",
]


def as_date(record):
    value = str(record.get("created") or record.get("updated") or "")
    return value[:10] if len(value) >= 10 else "1970-01-01"


def rewrite_table(sheet, records, headers):
    sheet.delete_rows(1, max(sheet.max_row, 1))
    sheet.append(headers)
    for record in records:
        sheet.append([record.get(header) for header in headers])


def rewrite_json_sheet(sheet, records):
    sheet.delete_rows(1, max(sheet.max_row, 1))
    sheet.append(["Date", "Data"])
    groups = defaultdict(list)
    for record in records:
        groups[as_date(record)].append(record)
    for date in sorted(groups):
        sheet.append([date, json.dumps(groups[date], ensure_ascii=False, separators=(",", ":"))])


def main():
    if len(sys.argv) != 3:
        raise SystemExit("Usage: sync_local_crm_to_xlsx.py LOCAL-CRM.json WORKBOOK.xlsx")
    json_path = Path(sys.argv[1])
    workbook_path = Path(sys.argv[2])
    data = json.loads(json_path.read_text(encoding="utf-8"))
    workbook = openpyxl.load_workbook(workbook_path)

    user_headers = [
        "Username", "Email", "Password", "Role", "Status", "ProfileImage",
        "ThemeMode", "CustomColors", "CreatedAt", "CreatedBy", "UpdatedAt",
        "UpdatedBy", "MonthlyTarget",
    ]
    role_headers = ["role_key", "label", "color", "sort_order", "is_super", "hidden_signup", "permissions"]
    roles = []
    for role in data.get("roles", []):
        item = dict(role)
        item["permissions"] = json.dumps(item.get("permissions") or {}, ensure_ascii=False, separators=(",", ":"))
        roles.append(item)
    rewrite_table(workbook["Users"], data.get("users", []), user_headers)
    rewrite_table(workbook["Roles"], roles, role_headers)
    for name in JSON_SHEETS:
        rewrite_json_sheet(workbook[name], data.get("sheets", {}).get(name, []))

    temp_path = workbook_path.with_suffix(".syncing.xlsx")
    workbook.save(temp_path)
    os.replace(temp_path, workbook_path)
    print(f"Synced local CRM data to {workbook_path}")


if __name__ == "__main__":
    main()
