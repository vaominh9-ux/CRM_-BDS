import json
import sys
from pathlib import Path

import openpyxl


def read_json_rows(sheet):
    records = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        raw = row[1] if len(row) > 1 else None
        if not raw:
            continue
        value = json.loads(raw)
        records.extend(value if isinstance(value, list) else [value])
    return records


def location_path_builder(locations):
    by_id = {item.get("id"): item for item in locations}

    def location_path(location_id):
        names = []
        current = by_id.get(location_id)
        seen = set()
        while current and current.get("id") not in seen:
            seen.add(current.get("id"))
            names.append(current.get("name", ""))
            current = by_id.get(current.get("parentId"))
        return " › ".join(reversed([name for name in names if name]))

    return location_path


def build_portal(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    locations = [item for item in read_json_rows(workbook["Locations"]) if not item.get("deleted")]
    amenities = [item for item in read_json_rows(workbook["Amenities"]) if not item.get("deleted")]
    properties = read_json_rows(workbook["Properties"])
    location_path = location_path_builder(locations)
    amenities_by_id = {
        item.get("id"): {"name": item.get("name", ""), "icon": item.get("icon", "")}
        for item in amenities
    }

    visible = []
    for item in properties:
        if item.get("deleted") or not item.get("publishedAt") or item.get("status") not in ("Available", "Reserved"):
            continue
        visible.append({
            "id": item.get("id"),
            "referenceCode": item.get("referenceCode", ""),
            "title": item.get("title", ""),
            "slug": item.get("slug", ""),
            "description": item.get("description", ""),
            "propertyType": item.get("propertyType", ""),
            "listingType": item.get("listingType", ""),
            "status": item.get("status", ""),
            "price": item.get("price", 0),
            "rentFrequency": item.get("rentFrequency", ""),
            "areaSize": item.get("areaSize", 0),
            "areaUnit": item.get("areaUnit", ""),
            "bedrooms": item.get("bedrooms"),
            "bathrooms": item.get("bathrooms"),
            "locationId": item.get("locationId"),
            "locationPath": location_path(item.get("locationId")),
            "address": item.get("address", ""),
            "latitude": item.get("latitude"),
            "longitude": item.get("longitude"),
            "isFeatured": 1 if item.get("isFeatured") else 0,
            "viewsCount": item.get("viewsCount", 0),
            "images": item.get("images") or [],
            "amenities": [amenities_by_id[amenity_id] for amenity_id in item.get("amenityIds", []) if amenity_id in amenities_by_id],
            "publishedAt": item.get("publishedAt"),
        })

    visible.sort(key=lambda item: item.get("publishedAt") or "")
    visible.reverse()
    return {
        "success": True,
        "properties": visible,
        "locations": [
            {"id": item.get("id"), "parentId": item.get("parentId"), "name": item.get("name", ""), "level": item.get("level", "")}
            for item in locations
        ],
        "amenities": [
            {"id": item.get("id"), "name": item.get("name", ""), "icon": item.get("icon", "")}
            for item in amenities
        ],
    }


def read_table(sheet):
    rows = list(sheet.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [str(value or "") for value in rows[0]]
    return [
        {headers[index]: value for index, value in enumerate(row) if index < len(headers)}
        for row in rows[1:]
        if any(value is not None for value in row)
    ]


def build_local_crm(workbook_path):
    workbook = openpyxl.load_workbook(workbook_path, read_only=True, data_only=True)
    json_sheets = ["Locations", "Amenities", "Properties", "Leads", "FollowUps", "Appointments", "Owners", "Deals", "Tenancies", "Logs"]
    roles = read_table(workbook["Roles"])
    for role in roles:
        try:
            role["permissions"] = json.loads(role.get("permissions") or "{}")
        except json.JSONDecodeError:
            role["permissions"] = {}
    return {
        "users": read_table(workbook["Users"]),
        "roles": roles,
        "sheets": {name: read_json_rows(workbook[name]) for name in json_sheets},
    }


def main():
    if len(sys.argv) not in (3, 4):
        raise SystemExit("Usage: export_xlsx_data.py INPUT.xlsx PORTAL.json [LOCAL-CRM.json]")
    input_path = Path(sys.argv[1])
    output_path = Path(sys.argv[2])
    output_path.parent.mkdir(parents=True, exist_ok=True)
    portal = build_portal(input_path)
    output_path.write_text(json.dumps(portal, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Exported {len(portal['properties'])} public properties to {output_path}")
    if len(sys.argv) == 4:
        local_output_path = Path(sys.argv[3])
        local_output_path.parent.mkdir(parents=True, exist_ok=True)
        local_crm = build_local_crm(input_path)
        local_output_path.write_text(json.dumps(local_crm, ensure_ascii=False, indent=2, default=str), encoding="utf-8")
        print(f"Exported local CRM data to {local_output_path}")


if __name__ == "__main__":
    main()
